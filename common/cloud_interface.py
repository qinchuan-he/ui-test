#coding=utf-8

from common.comfunction import send_mail
import os
from common.comfunction import OpenBrowser
from time import sleep
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from common.private import EmailProperty,folder_path,DB
import shutil
import py
import openpyxl
from typing import List
import requests

def insert_content(parameter:List):
    """ 往数据库插入东西"""
    url = r'D:\5python'
    host1 = DB.host
    port1 = DB.port
    user1 = DB.user
    password1 = DB.password
    db_name = DB.db_name
    connect = pymysql.connect(host =host1,port = int(port1),user = user1,password=password1,database='test',charset='utf8')
    cur = connect.cursor()
    # s = cur.execute('select * from consumer')
    #     # print(cur.fetchall())
    try:
        sql = "insert into ebook_name(path,`name`) values(%s,%s);"
        print(sql)
        # ls = [('4','3','4'),('3','3','4')]
        # cur.execute("insert into ebook_name(path,`name`,`desc`) values('2','3','4');") # 单条方法
        cur.executemany(sql,parameter) # 批量方法
        connect.commit()
    except Exception as e:
        connect.rollback()
        print('出现异常，终止')
        print(e)
    print('执行完毕')

def read_excel(position=None):
    """ 读取Excel,position有值就全部循环"""
    url = r'D:\1212.xlsx'
    url = r'D:\5python\800G小说目录.xlsx'
    url = r'D:\work\1测试\1需求\cyprex2.1.9'
    wk = openpyxl.open(url)
    all = []
    for i in wk.worksheets:
        count = i.max_row
        for j in range(count):
            path = i.cell(j+1,1).value
            name = i.cell(j+1,2).value
            ll = (path,name)
            all.append(ll)
            print('-')
            # if j == 20:
            #     break

        # if not position:
        #     break
    return all

import hashlib
import json
import base64
def generate_jwt(**kwargs):
    """调试jwt的测试接口"""
    user_open_id = kwargs.get("user_open_id", "680")
    username = kwargs.get("username", "员工")

    p = hashlib.sha256('MzY1Mjk0OTM5NTgz'.encode("utf-8"))
    header = {
        "alg": "HS256", # 加密方式为HMAC SHA256
        "typ": "JWT"
    }
    header = base64.urlsafe_b64encode(json.dumps(header).encode('utf-8')).decode('utf-8')
    payload = {
        "iat": time.time(),  # 加密方式为HMAC SHA256
        "exp": time.time() + 60 * 60,
        "user_open_id": user_open_id,
        "username": username,
    }
    payload = base64.urlsafe_b64encode(json.dumps(payload).encode('utf-8')).decode('utf-8')
    p.update(r"{}\.{}".format(header, payload).encode("utf-8"))
    signature = p.hexdigest()

    return "Bearer {}.{}.{}".format(header, payload, signature)

def interface(head_key,resource_id=None):
    """  紫光云接口"""
    header = {"Authorization":head_key}
    print(header)
    cookie = {'fir_session_id':'vkbbijcvul6y8hw7kkzvoqp2jvbw9bv2'}

    team_type= '3'
    team_open_id = '222225'
    # 创建会议接口
    # url = 'https://testcloud.fir.ai/api/group/meeting/create/'
    # # url = 'https://devcloud.fir.ai/api/group/meeting/create/'
    # # url = 'http://117.9.200.62:20080/api/group/meeting/create/'
    # parameters = {'name':'IM管理','team_type':team_type,'team_open_id':team_open_id,'extra_info':'来自接口'}
    # res = requests.post(url=url,data=parameters,headers=header)
    # print(res.text)

    # 添加会议成员
    # url='https://testcloud.fir.ai/api/group/meeting/member/create/'
    # pp = [{'user_open_id':'1','user_name':'气喘','name':'hahah'},{'user_open_id':'33','user_name':'气喘','name':'hahah'}]
    # parameter={'team_type':'3','team_open_id':'222224','user_open_id_list':json.dumps(pp)}
    # res = requests.post(url=url,data=parameter,headers = header)
    # print(res.text)

    # s上传会议文件
    url = 'https://testcloud.fir.ai/api/group/meeting/resource/upload/'
    # url = 'http://117.9.200.62:20080/api/group/meeting/resource/upload/'
    # file_path =os.path.join(r'D:\上传文件\自动化验证文档\19种格式\office')
    # file_path = os.path.join(r'D:\上传文件\自动化验证文档\19种格式\图例提取文件')
    # file_path = os.path.join(r'D:\上传文件\自动化验证文档\19种格式\合并')
    # file_path = os.path.join(r'D:\上传文件\自动化验证文档\19种格式\比对文件')
    file_path = r'D:\上传文件\自动化验证文档\19种格式\图片'
    file_path = r'D:\上传文件\自动化验证文档\19种格式\其他'
    count = 430
    for i in os.listdir(file_path):
        print(i)
        file_info  = {'name':i,'author':'qinchuan'}
        file={'file':open(os.path.join(file_path,i),'rb')}
        date = {'team_type':team_type,'team_open_id':team_open_id,'resource_id':count,'file_info':json.dumps(file_info)}
        res = requests.post(url=url,data=date,headers = header,files=file)
        print(res.text)
        print(count)
        count+=1

    # 文件重命名
    # url = 'https://testcloud.fir.ai/api/group/meeting/resource/update/'
    # date = {'resource_id':'50','name':'单一名字.pdf'}
    # res = requests.post(url=url,data=date,headers=header)
    # print(res.text)

    # 文件删除
    # url='https://testcloud.fir.ai/api/group/meeting/resource/delete/'
    # data = {'resource_id':'50'}
    # res = requests.post(url=url,data=data,headers=header)
    # print(res.text)




    # 接口
    # url = 'https://testcloud.fir.ai/api/resource/data/fetch/'
    # resource_id = 'AxrEkReGG4rP52BY'
    # request_url=url+'?id='+resource_id+'&view_type=preview'
    # print(request_url)
    # url_s='https://testcloud.fir.ai/api/resource/data/fetch/?id=AxrEkReGG4rP52BY&view_type=preview'
    # res = requests.get(url=request_url,headers = header)
    # print(res.text)



if __name__=='__main__':
    # sendemail1()
    # yousee()
    # insert_content()
    # result = read_excel()
    # insert_content(result)
    key = generate_jwt()
    # print(key)
    interface(key)  # 会议系统和IM系统接口









































