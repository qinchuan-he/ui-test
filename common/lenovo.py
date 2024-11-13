import hmac
import json
import base64
import datetime
import uuid

import requests
import openpyxl


def gen(url, params):
    """user_id int类型  rev 版本号 不传默认为打开最新版本"""

    temp = base64.b64encode(json.dumps(params).encode("utf8"))
    api_key = '49aea739-dedf-4ad4-9a65-8356f715b3db'
    sign = hmac.new(api_key.encode('utf8'), msg=temp, digestmod="MD5").hexdigest()
    params = temp.decode()
    # print(params)
    # print(sign)

    url = "%sparams=%s&sign=%s" % (url, params, sign)
    # print(url)
    return url

def test_api(file_name,file_id,rev,user_id,view_type):
    """ 测试环境使用"""
    url = "http://192.168.1.224:8061/api/file/parse?"
    params = {"file_id": file_id, "rev": rev, "opt_type": "01"}
    url_3 = gen(url, params)
    res = requests.get(url_3)
    print(res.text)
    url = "http://192.168.1.224:8060/data/fetch?"
    params = {"file_id": file_id, "view_type": view_type, "file_name": file_name, "rev": rev, "user_id": user_id}
    gen(url, params)

# 读取Excel文档获取neid和rev
def read_excel():
    rest = []
    wk = openpyxl.load_workbook('lenovo.xlsx')
    wb = wk.worksheets[0]
    print(wb.cell(2,2).value)
    row = wb.max_row
    for i in range(2,row+1):
        a = {'rev':wb.cell(i,2).value,'neid':wb.cell(i,3).value}
        rest.append(a)
        # print(a)
    return  rest

# 创建联想网盘的更新索引接口url，存入文件
def create_url(lst):
    url = 'http://192.168.1.224:8061/api/file/parse?'
    url = 'api/file/parse?'
    file = open(r'D:\work\1测试\16测试数据\1.txt', 'w+')
    for i in lst:
        params = {"file_id": i.get('neid'), "rev": i.get('rev'), "opt_type": "01"}
        url_2 = gen(url, params)
        # print(url_2)
        file.write(url_2+str('\n'))

# 造预览调用接口url
def create_url2(lst):
    url = 'http://192.168.1.224:8060/data/fetch?'
    url =  '/api/data/fetch?'
    file = open(r'D:\work\1测试\16测试数据\preview.txt','w+')
    for i in lst:
        params = {"view_type":"preview","file_id": i.get('neid'), "file_name":"cesces.pdf","rev": i.get('rev'),"user_id":27}
        url_2 = gen(url,params)
        file.write(url_2+str('\n'))
    file.close()


if __name__ == '__main__':
    # 必传rev
    # url = "http://localhost:4000/data/fetch?"
    # params = {"file_id": 141, "view_type": "preview", "file_name": "新建 Microsoft Word 文档.docx", "rev": "6dc8564a15644c8aae70774f25b5eb01", "user_id": 27}

    # 第一入口解析
    #     # url = "http://192.168.1.62:8099/api/file/parse?"
    #     # params = {"file_id": 141,  "rev": "6dc8564a15644c8aae70774f25b5eb01", "opt_type": "01"}

    # 测试环境,先调数据组接口，再生成前端页面访问接口
    # user_id = 27
    # file_name = '图例验证文件.pdf'
    # neid = 144
    # rev = '129e5cf528ea4857b11c03982a52d770'
    user_id = 27
    file_name = '表格图片.doc'
    neid = 145
    rev = '0c1466f88a124d32bb244eccc6a39217'
    view_type = 'preview'
    # view_type = 'editor'
    test_api(file_name,neid,rev,user_id,view_type)

    # res = read_excel()
    # create_url(res) # 造索引，
    # create_url2(res) # 预览




