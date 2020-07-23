#coding:utf-8
import requests
import os
import threading
import time
import re
import openpyxl


def dd(file_url=None):
    if not file_url:
        file_url = r'D:\上传文件\pdf比对\投标样本\华尔街\公告\天齐锂业股东大会决议.PDF'
    if type(file_url)==str:
        file_name = os.path.split(file_url)[1]
        url = 'http://124.77.120.212:10151/v2/files/databox/第一个文件夹/'+file_name+'?X-LENOVO-SESS-ID=e84htk5ps920sa8g7jtv49i5d5&S=D03F2ADD&uid=27&overwrite=true&source=file&language=zh&t=&path_type=ent&1593506087757'
        # cookie = {'X-LENOVO-SESS-ID':'e84htk5ps920sa8g7jtv49i5d5'}
        url_2 = 'http://124.77.120.212:10151/v2/files/databox/第二个文件夹/'+file_name+'?X-LENOVO-SESS-ID=e84htk5ps920sa8g7jtv49i5d5&S=D03F2ADD&uid=27&overwrite=true&source=file&language=zh&t=&path_type=ent&1593506087757'
        url_3 = 'http://124.77.120.212:10151/v2/files/databox/第三个文件夹/'+file_name+'?X-LENOVO-SESS-ID=e84htk5ps920sa8g7jtv49i5d5&S=D03F2ADD&uid=27&overwrite=true&source=file&language=zh&t=&path_type=ent&1593506087757'

        file = {'file':open(file_url,'rb')}
        file_2 = {'file': open(file_url, 'rb')}
        file_3 = {'file': open(file_url, 'rb')}
        try:
            res = requests.post(url=url,files = file)
            res_2 = requests.post(url=url_2, files=file_2)
            res_3 = requests.post(url=url_3, files=file_3)
            # print(res.text)
            # print(res_2.text)
            # print(res_3.text)
        except Exception as e:
            print(e)
    else:
        for f_url in file_url:
            file_name = os.path.split(f_url)[1]
            url = 'http://124.77.120.212:10151/v2/files/databox/第一个文件夹/' + file_name + '?X-LENOVO-SESS-ID=e84htk5ps920sa8g7jtv49i5d5&S=D03F2ADD&uid=27&overwrite=true&source=file&language=zh&t=&path_type=ent&1593506087757'
            # cookie = {'X-LENOVO-SESS-ID':'e84htk5ps920sa8g7jtv49i5d5'}
            url_2 = 'http://124.77.120.212:10151/v2/files/databox/第二个文件夹/' + file_name + '?X-LENOVO-SESS-ID=e84htk5ps920sa8g7jtv49i5d5&S=D03F2ADD&uid=27&overwrite=true&source=file&language=zh&t=&path_type=ent&1593506087757'
            url_3 = 'http://124.77.120.212:10151/v2/files/databox/第三个文件夹/' + file_name + '?X-LENOVO-SESS-ID=e84htk5ps920sa8g7jtv49i5d5&S=D03F2ADD&uid=27&overwrite=true&source=file&language=zh&t=&path_type=ent&1593506087757'

            file = {'file': open(f_url, 'rb')}
            file_2 = {'file': open(f_url, 'rb')}
            file_3 = {'file': open(f_url, 'rb')}
            try:
                res = requests.post(url=url, files=file)
                res_2 = requests.post(url=url_2, files=file_2)
                res_3 = requests.post(url=url_3, files=file_3)
                # print(res.text)
                # print(res_2.text)
                # print(res_3.text)
            except Exception as e:
                print(e)

class Mythread(threading.Thread):

    def __init__(self,file_url):
        threading.Thread.__init__(self)
        self.file_url = file_url
    def run(self) -> None:
        dd(self.file_url)

def driver():
    root_path = os.path.dirname(os.path.abspath(__file__))
    # root_path = '/data/judge'
    lst = os.listdir(root_path)
    num = 1
    ls = []
    date1 = time.time()
    for i in lst:
        if os.path.splitext(i)[1]=='.pdf'or os.path.splitext(i)[1]=='.PDF':
            # print(i)
            ls.append(os.path.join(root_path,i))
    sp = int(len(ls)/5)
    a1 = Mythread(ls[0:sp])
    a2 = Mythread(ls[sp:2 * sp])
    a3 = Mythread(ls[2 * sp:3*sp])
    a4 = Mythread(ls[3 * sp:4*sp])
    a5 = Mythread(ls[4 * sp:])
    a1.start()
    a2.start()
    a3.start()
    a4.start()
    a5.start()

    a1.join()
    a2.join()
    a3.join()
    a4.join()
    a5.join()
    print('------------执行完成-------------')
    print('总计耗时：%s 秒'%(int(time.time())-int(date1)))

# 获取联想网盘中文件id
def get_fileid():
    """"""
    rev_list = []
    neid_list = []
    res_list = []

    histoty_result = {'neid':0}
    cookie = {'X-LENOVO-SESS-ID': 'e84htk5ps920sa8g7jtv49i5d5'}
    folders = ['第一个文件夹','第二个文件夹','第三个文件夹']
    for folder in folders:
        offset=0
        count_offset=0
        for i in range(0,1520):
        # i = 1512
            if count_offset%100==0:
                offset=count_offset
            url = 'http://124.77.120.212:10150/v2/metadata_page/databox/'+folder+'?path_type=ent&page_button_count=5' \
                  '&include_deleted=false&page_size=20&waitContent=.lui-filelist&page_num='+str(i)+'&offset='+str(offset)+'&sort=desc&orderby=mtime' \
                  '&_=1593570918608&account_id=1&uid=27&S=D03F2ADD'

            res = requests.get(url,cookies = cookie)
            # print(res.text)
            count_offset=(i+1)*20
            neid = re.findall('"neid":(.*?),', res.text)
            rev = re.findall('"rev":"(.*?)"', res.text)
            for j in range(len(rev)):
                if int(neid[j+1]) in neid_list:
                    print('出现重复，请求page_num：%d  重复neid：%s'%(i,neid[j+1]))
                    # print(neid_list)
                    continue
                else:
                    neid_list.append(int(neid[j+1]))
                key = {}
                key['rev'] = rev[j]
                key['neid'] = neid[j + 1]
                res_list.append(key)

    # print(res_list)
    return res_list

# 内容写入Excel
def witer_excel(res_list=None):
    wb = openpyxl.Workbook()
    ws =wb.active
    ws.cell(1,1).value='序号'
    ws.cell(1,2).value='rev'
    ws.cell(1,3).value='neid'
    row = 2
    # res_list = [{'rev': 'dc48275c23634a78b8dd82d161de9698', 'neid': '90692'}, {'rev': '711d019455dd40359ab6eeff062df2fb', 'neid': '90687'}, {'rev': '6294a1c6ca5a4440a4613b9d231df486', 'neid': '90686'}, {'rev': 'f18d06112fb1411a848fd3e25c4a733d', 'neid': '90681'}, {'rev': '8434d293fabe44cab7d8697d594b5c12', 'neid': '90680'}, {'rev': 'b4c6e4a298874f73b13e03a09d9e4c5b', 'neid': '90675'}, {'rev': 'de7cbe051e5945ef9c96f12dc3d5e62e', 'neid': '90674'}, {'rev': '9902bfba87fc49abadaf1d6ecdfb9603', 'neid': '90668'}, {'rev': 'b42c2915a9054ad68870089513888ec0', 'neid': '90669'}, {'rev': 'bdd13dde7f424b679942c42d4b478afd', 'neid': '90662'}, {'rev': '5ae7522eb2ed4bf2b67bc43f95f0209e', 'neid': '90663'}, {'rev': '8f77e052ca14453bbfddf48356d6dd7c', 'neid': '90657'}, {'rev': '5cc771424d704044b05f5000dacc7936', 'neid': '90655'}, {'rev': 'd59e77d31a7f40579bbbc445c0732bab', 'neid': '90650'}, {'rev': 'd909ff911af44ad5b3137501a89da74d', 'neid': '90651'}, {'rev': 'e7927ab935324227bac509dd6822fc7f', 'neid': '90645'}, {'rev': '18c67c7ff4444c558c778880d5a89644', 'neid': '90644'}, {'rev': 'd506e52c31a34040ba40e97e37c1c6c4', 'neid': '90637'}, {'rev': '8c03b085fcc2429f96d50b9a7f9c7fc3', 'neid': '90640'}, {'rev': '7678997063c142158fc405bcd9a99b60', 'neid': '90633'}]

    if res_list:
        for i in res_list:
            ws.cell(row,1).value=row-1
            ws.cell(row,2).value=i.get('rev')
            ws.cell(row,3).value=i.get('neid')
            row+=1

        wb.save(r'D:\上传文件\office文件\难点文件\lenovo.xlsx')
        print('写入完成')

def cc():
    res_list = [{'rev': '711d019455dd40359ab6eeff062df2fb', 'neid': '90687'}, {'rev': '711d019455dd40359ab6eeff062df2fb', 'neid': '90687'}, {'rev': '6294a1c6ca5a4440a4613b9d231df486', 'neid': '90686'}, {'rev': 'f18d06112fb1411a848fd3e25c4a733d', 'neid': '90681'}, {'rev': '8434d293fabe44cab7d8697d594b5c12', 'neid': '90680'}, {'rev': 'b4c6e4a298874f73b13e03a09d9e4c5b', 'neid': '90675'}, {'rev': 'de7cbe051e5945ef9c96f12dc3d5e62e', 'neid': '90674'}, {'rev': '9902bfba87fc49abadaf1d6ecdfb9603', 'neid': '90668'}, {'rev': 'b42c2915a9054ad68870089513888ec0', 'neid': '90669'}, {'rev': 'bdd13dde7f424b679942c42d4b478afd', 'neid': '90662'}, {'rev': '5ae7522eb2ed4bf2b67bc43f95f0209e', 'neid': '90663'}, {'rev': '8f77e052ca14453bbfddf48356d6dd7c', 'neid': '90657'}, {'rev': '5cc771424d704044b05f5000dacc7936', 'neid': '90655'}, {'rev': 'd59e77d31a7f40579bbbc445c0732bab', 'neid': '90650'}, {'rev': 'd909ff911af44ad5b3137501a89da74d', 'neid': '90651'}, {'rev': 'e7927ab935324227bac509dd6822fc7f', 'neid': '90645'}, {'rev': '18c67c7ff4444c558c778880d5a89644', 'neid': '90644'}, {'rev': 'd506e52c31a34040ba40e97e37c1c6c4', 'neid': '90637'}, {'rev': '8c03b085fcc2429f96d50b9a7f9c7fc3', 'neid': '90640'}, {'rev': '7678997063c142158fc405bcd9a99b60', 'neid': '90633'}]
    # res_list = [1,1,2,3]
    print(len(res_list))
    res_list = list(set(res_list))
    print(len(res_list))

if __name__=="__main__":
    # dd()
    # driver()
    # cc()
    res_list = get_fileid()
    witer_excel(res_list)