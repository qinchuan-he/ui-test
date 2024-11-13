# -*- coding:utf-8 -*-

import os
import json
import requests
import re
import fitz
import openpyxl
import time
import threading

# 公告和研报解析接口,写入Excel
def analysis(folder_path = None, style=None,in_lst=None,single=None):
    """ 数据组pdf公告解析接口"""
    if not folder_path:
        # folder_path = 'D:\\上传文件\\pdf比对\\数据组--样本文件\\研报' # 研报数据
        folder_path = r'D:\上传文件\pdf比对\数据组--样本文件'  # 公告数据
        # folder_path = '/data/judge/'
        folder_path = r'D:\上传文件\pdf比对\投标样本\华尔街\公告'

    url = 'http://192.168.1.219:8016/data/api/algorithm/algorithm'
    url = 'http://124.77.120.212:5016/data/api/algorithm/algorithm'
    # argument = {'algo_type':'pdf2html','pdf_type':'notice'}   # 公告
    # argument = {'algo_type':'pdf2html','pdf_type':'report'} # 研报
    argument = {'algo_type': 'excel_parser'}  # Excel
    # argument = {'algo_type': 'word_parser'}  # word

    res_url_part = 'http://124.77.120.212:5016/data/api/store/download/?full_path='
    if style=='excel':
        suffix1 = '.xls'
        suffix2 = '.xlsx'
    elif style=='word':
        suffix1 = '.doc'
        suffix2 = '.docx'
    else:
        suffix1 = '.pdf'
        suffix2 = '.PDF'
    excel_url = os.path.join(folder_path,time.strftime('%y-%m-%d_%H_%M_%S',time.localtime(time.time()))+'解析结果.xlsx')

    if in_lst:
        path = in_lst
    else:
        path = os.listdir(folder_path)
    lst = []
    # for i in os.listdir(os.path.join(folder_path)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1,1).value='序号'
    ws.cell(1,2).value = '文件名'
    ws.cell(1,3).value = '地址'
    row = 2
    count = 1
    for i in path:
        print('计数：%d'%count)
        count+=1
        if os.path.splitext(i)[1]==suffix1 or os.path.splitext(i)[1]==suffix2:
            file_path = os.path.join(folder_path,i)
            file = {'file': open(file_path, 'rb')}
            res = requests.post(url,data=argument,files=file)
            if res.status_code==200 and res.json().get('status')==1:
                res_text=res.text
                ws.cell(row, 1).value = row - 1
                ws.cell(row, 2).value = i
                try:
                    if style=='excel' or style=='word':
                        res_j = res_text
                        url_value = re.findall('"uid":"(.*?)"', res_j)[0]
                        url_value = res_url_part+url_value+'.html'
                    else:
                        res_j = json.loads(res_text).get('data')
                        url_value = re.findall("href='(.*?)'",res_j)[0]

                    # ws.cell(row,1).value=row-1
                    # ws.cell(row,2).value=i
                    ws.cell(row, 3).value = url_value
                    # row += 1
                    result_single = {'name':i,'url':url_value}

                except Exception as e:
                    ws.cell(row, 3).value = str(e)
                    result_single = {'name':i,'url': '获取地址出问题: %s'%e}
                    print('获取地址出问题: %s'%e)
            else:

                    print('解析失败了。。。')
                    result_single={'name':i,'url': '这个解析失败了: %s'%res.text}


            lst.append(result_single)
            row += 1
    if not single:
        wb.save(excel_url)
        print('---------------执行完成----------------')
    return lst

# 计算解析页码
def total_page(url_s,file_s,file_path):
    """ 传入单个str或者数组"""
    urls=[]
    files = []
    if type(url_s)==str:
        urls.append(url_s)
        files.append(file_s)
    else:
        urls = url_s
        files = file_s
    for i in range(len(urls)):
        res = requests.get(urls[i])
        # print(res.text)
        # print(res.text.encode('utf8'))
        rej = re.findall('/section',res.text)
        # print('HTML格式页数为： %d'%len(rej))
        try:
            with open(os.path.join(file_path,files[i]), "rb") as pdf_file:
                pdf_doc_fitz = fitz.open(filetype="pdf", stream=pdf_file.read())
                # print("\t总页数%d" % pdf_doc_fitz.pageCount)
        except Exception as e:
            print("fitz加载失败！ %s" % e.__str__())
        if len(rej)==pdf_doc_fitz.pageCount:
            # print('%s %s'%(files[i],'解析未丢页'))
            pass
        else:
            print('%s ------- %s        html页数:原文页数=%s:%s'%(files[i],'解析丢页',len(rej),pdf_doc_fitz.pageCount))

# 读取Excel
def read_excel(url = None):
    """ """
    url = r'D:\上传文件\pdf比对\投标样本\华尔街\公告\20-06-30_11_31_31解析结果.xlsx'
    wb = openpyxl.load_workbook(url)
    content = wb.worksheets[0]
    count = content.max_row
    # print(content.cell(2,2).value)
    # print(count)
    file_names = []
    file_paths = []
    for i in range(2,count+1):
        file_names.append(content.cell(i,2).value)
        file_paths.append(content.cell(i,3).value)
    # print(file_names)
    # print(file_paths)
    return file_names,file_paths

# 多线程类,启动解析
class mythreads(threading.Thread):
        result = []
        def __init__(self,folder_path,style,in_lst,single):
            threading.Thread.__init__(self)
            self.folder_path = folder_path
            self.style = style
            self.in_lst = in_lst
            self.single = single
        def run(self):
            self.result = analysis(self.folder_path,self.style,self.in_lst,self.single)


def threads(folder_path,style):
    in_list = os.listdir(folder_path)

    sp = int(len(in_list)/4)
    single = 'thread'
    a1 = mythreads(folder_path,style,in_list[0:sp],single)
    a2 = mythreads(folder_path, style, in_list[sp:2*sp],single)
    a3 = mythreads(folder_path, style, in_list[2*sp:3*sp],single)
    a4 = mythreads(folder_path, style, in_list[3*sp:],single)

    a1.start()
    a2.start()
    a3.start()
    a4.start()

    a1.join()
    a2.join()
    a3.join()
    a4.join()

    print('------------------執行完成')

    # 创建Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1).value = '序号'
    ws.cell(1, 2).value = '文件名'
    ws.cell(1, 3).value = '地址'
    row = 2
    res_s = []
    res_s.append(a1.result)
    res_s.append(a2.result)
    res_s.append(a3.result)
    res_s.append(a4.result)
    for j in res_s:
        for i in range(len(j)):
            ws.cell(row, 1).value = row - 1
            ws.cell(row, 2).value = j[i].get('name')
            ws.cell(row, 3).value = j[i].get('url')
            row+=1
    excel_url = os.path.join(folder_path, time.strftime('%Y-%m-%d_%H_%M_%S', time.localtime(time.time())) + '解析结果.xlsx')
    wb.save(excel_url)





if  __name__=='__main__':
    # analysis()   # 数据组公告接口
    # url = r'http://124.77.120.212:9650/apis/download/?data_id=2020-06-29/7d2a24e0-b9b2-11ea-b925-0242ac110008'
    # file_path = r'D:\上传文件\pdf比对\投标样本\华尔街\公告'
    # file_name = '申万宏源十大金股组合.pdf'
    # path = 'D:\上传文件\pdf比对\投标样本\华尔街\公告'
    # file_name,file_path=read_excel()
    # total_page(file_path,file_name,path) # 统计解析页数
    # # read_excel()
    folder_path = r'D:\上传文件\office文件\Excel文件\测试解析用'
    style = 'excel'
    threads(folder_path,style)
    # style1 = 'word'
    # folder_path = r'D:\上传文件\office文件\Word文件'
    # threads(folder_path,style1)  # 启用多线程的执行
