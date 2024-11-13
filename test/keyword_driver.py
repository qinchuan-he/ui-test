

import openpyxl
from openpyxl.styles import PatternFill
from common.comfunction import *
from time import sleep
from test.keyword_report import createReportHtml




#增加关键字驱动模块
# 1.读取（xlsx格式），进行轮询取出每一行
# 2.每一行根据关键字执行

def open():
    """ test Excel执行"""
    # wb = openpyxl.load_workbook('D:\\work\\1测试\\7设计\\关键字对照表.xlsx')
    file = 'D:\\work\\1测试\\2用例\\cypress系统\\回归用例\\私有文件_新建文件_用例.xlsx'
    file = 'D:\\work\\1测试\\2用例\\cypress系统\\回归用例\\私有_新建文件夹_笔记_脑图_白板.xlsx'
    file = 'D:\\work\\1测试\\2用例\\cypress系统\\回归用例\\测试用.xlsx'

    exec_path = os.path.join(os.path.dirname(file),'执行结果')
    report_path = os.path.join(os.path.dirname(file),'执行结果\\'+time.strftime('%Y_%m_%d_%H_%M_%S',time.localtime(time.time()))+'report.html')
    # picture_path = os.path.join(os.path.dirname(file),'执行结果\\'+time.strftime('%Y_%m_%d_%H_%M_%S',time.localtime(time.time()))+'.png')
    if not os.path.exists(exec_path):
        os.mkdir(exec_path)
    file_path = os.path.abspath(file)  # 获取文件绝对路径
    file_split = os.path.splitext(os.path.basename(file_path))  # 获取文件名（不带后缀）
    wb = openpyxl.load_workbook(os.path.abspath(file_path))  # 读取Excel文件
    sheet = wb.worksheets
    all = []
    result = []  # 存放执行结果
    for i in sheet:
        start_time = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
        result_num = 0
        j = i.max_row
        k = i.max_column
        k = 7  # 2023-01-31增加，列数写死
        # 清除执行时间和执行结果
        # if i != sheet[2]:
        #     continue
        for q in range(2,j+2):
            i.cell(q,7).value=''
        for row in range(2,j+1):
            rows = []
            column_s=''
            for column in range(1,k+1):
                cell = i.cell(row,column).value
                column_s = column
                rows.append(cell)
            if rows[2] and rows[3] and rows[4]:
                s = str(rows[2])+'('+'"'+str(rows[3])+'"'+','+'"'+str(rows[4])+'"'+')'
            elif rows[2] and rows[3]:
                s = str(rows[2])+'('+'"'+str(rows[3])+'"'+')'
            elif rows[2] and rows[4]:
                s = str(rows[2])+'('+str(rows[4])+')'
            elif rows[2]:
                s = str(rows[2])+'('+')'
            else:
                s = ''
            print(s)
            all.append(s)
            try:
                OpenBrowser()
                eval(s)  # eval是内置函数，执行字符串表达式。这里用来执行用例
                i.cell(row, column_s-1).value = time.strftime('%Y_%m_%d %H:%M:%S',time.localtime(time.time()))
                fill_color = PatternFill(fgColor='1E90FF',fill_type='solid')
                i.cell(row,column_s).value = '通过'
                i.cell(row, column_s).fill = fill_color
            except Exception as e:
                print('执行出错，终止')
                # 添加截图，知道位置
                picture_path = os.path.join(os.path.dirname(file),'执行结果\\'+time.strftime('%Y_%m_%d_%H_%M_%S',time.localtime(time.time()))+'.png')
                screenShot(url=picture_path)
                i.cell(row, column_s - 1).value = time.strftime('%Y_%m_%d %H:%M:%S', time.localtime(time.time()))
                i.cell(row,column_s).value = '失败'
                fill_color = PatternFill(fgColor='FF3E96',fill_type='solid')
                i.cell(row, column_s).fill = fill_color
                print(e)
                CloseBrowsers() # 执行出错，终止并关闭浏览器
                result_num = 1
                end_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
                result_part = [i.title, '执行失败', start_time, end_time,e,picture_path]
                result.append(result_part)
                break
        end_time = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
        if result_num==0:
            result_part = [i.title, '执行成功', start_time, end_time]
            result.append(result_part)

    print(all)
    print('执行完毕')
    date = time.strftime('%Y_%m_%d_%H_%M_%S',time.localtime(time.time()))
    wb.save(os.path.join(exec_path,file_split[0]+'_执行_'+date+file_split[1]))
    print(result)
    # 增加生成HTML执行结果
    createReportHtml(report_path,result)




if __name__=='__main__':
    open()
































