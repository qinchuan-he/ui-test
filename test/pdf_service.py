#coding:utf-8

# pdf 解析服务相关

import requests
import os
import openpyxl
import threading

# exception_notice = 'D:\\work\\1测试\\15数据组\\公告研报分类\\公告.txt'
# exception_report = 'D:\\work\\1测试\\15数据组\\公告研报分类\\研报.txt'
# res_path = os.path.join(r"D:\work\1测试\15数据组\公告研报分类\结果.txt")
# res_path_report = os.path.join("D:\\work\\1测试\\15数据组\\公告研报分类\\结果是研报.txt")
# res_path_notice = os.path.join("D:\\work\\1测试\\15数据组\\公告研报分类\\结果是公告.txt")
# res_path_other = os.path.join("D:\\work\\1测试\\15数据组\\公告研报分类\\结果是其他.txt")

url_path = os.path.dirname(__file__)
exception_notice = os.path.join(url_path,'公告.txt')
exception_report = os.path.join(url_path,'研报.txt')
res_path = os.path.join(url_path,'结果.txt')
res_path_report = os.path.join(url_path,'结果是研报.txt')
res_path_notice = os.path.join(url_path,'结果是公告.txt')
res_path_other = os.path.join(url_path,'结果是其他.txt')

# 分类器
def judge_files(file_path=None):
    """ 判断PDF是研报还是公告"""
    # e_notice = open(exception_notice,'r',encoding='utf-8')
    # e_report = open(exception_report,'r',encoding='utf-8')
    # e_notices = e_notice.read()
    # e_reports = e_report.read()

    if not file_path:
        file_path=url_path
    url = "http://192.168.1.222/convert/pdf/pdf_fragment/pdf_clf"


    # names = os.listdir(os.path.join(file_path))
    names = os.listdir(os.path.dirname(__file__))
    print(type(names))
    print(names)

    res_file = open(res_path,'w+')
    res_file_report = open(res_path_report, 'w+')
    res_file_notice = open(res_path_notice, 'w+')
    res_file_other = open(res_path_other, 'w+')
    sum = 0
    fail = 0
    notice = 0
    report = 0
    other = 0
    for file_name in names:
        # if file_name not in e_notices and file_name not in e_reports:
        if os.path.splitext(file_name)[1]=='.pdf' or os.path.splitext(file_name)[1]=='.PDF':
            sum +=1
            file = os.path.join(file_path,file_name)
            files = {'file':open(file,'rb')}
            try:
                res = requests.post(url=url, files=files)
                # print(res.text)
                result = res.json()

                res_file.write(file_name+'  '+str(result.get('label')+'   '+str(result.get('confidence'))+result.get('msg')+'\n'))
                if result.get('msg')=='公告':
                    notice+=1
                    res_file_notice.write(file_name + '  ' + str(
                        result.get('label') + '   ' + str(result.get('confidence')) + result.get('msg') + '\n'))
                elif result.get('msg')=='研报':
                    report+=1
                    res_file_report.write(file_name + '  ' + str(
                        result.get('label') + '   ' + str(result.get('confidence')) + result.get('msg') + '\n'))
                else:
                    other+=1
                    res_file_other.write(file_name + '  ' + str(
                        result.get('label') + '   ' + str(result.get('confidence')) + result.get('msg') + '\n'))
                    print('识别出错')
            except Exception as e:
                fail+=1
                print('执行出错: %s'%file_name)
                print(e)
        # else:
        #     print('样本文件-剔除')

    res_file.close()
    res_file_notice.close()
    res_file_report.close()
    res_file_other.close()

    print('执行总共%s条，成功 %s条，失败%s条。其中公告 %s个，研报 %s个,其他类型 %s个'%(sum,notice+report+other,fail,notice,report,other))

def getUrl():
    # file_path = "D:\\上传文件\\pdf比对\\数据组--公告\\比亚迪：第六届董事会第二十次会议决议公告.pdf"
    file_path = "D:\\上传文件\\pdf比对\\数据组--公告"
    file_path = "D:\\上传文件\\pdf比对\\复杂文件"
    file_path = "D:\\上传文件\\pdf比对\\复杂文件\\非研报类" # 有报错
    file_path = "D:\\上传文件\\pdf比对\\复杂文件\\非研报类\\解析失败"
    file_path = "D:\\上传文件\\pdf比对\\1专业数据"   # 华为年报认错了
    file_path = "D:\\上传文件\\pdf比对\\1难点文件"
    file_path = "D:\\上传文件\\pdf比对\\3验证接口\\解析接口\\研报"
    file_path = "D:\\上传文件\\pdf比对\\3验证接口\\解析接口\\公告"
    file_path = "D:\\上传文件\\pdf比对\\3验证接口"
    file_path = "D:\\上传文件\\pdf比对\\1专业数据\\数据组-研报\\12\\1" # 样本文件-剔除
    file_path = "D:\\上传文件\\pdf比对\\1专业数据\\数据组-研报\\12\\7"  # 非pdf
    file_path = "D:\\上传文件\\pdf比对\\1专业数据\\数据组-研报\\12\\10"  # 样本文件-剔除
    file_path = "D:\\上传文件\\pdf比对\\1专业数据\\数据组-研报\\12\\21"  # 样本文件-剔除
    file_path = "D:\上传文件\pdf比对\繁体字文档"  # 样本文件-剔除,增加了replace方法，这里不用人工加\
    file_path = 'D:\上传文件\pdf比对\1专业数据\数据组-研报\2\1'
    file_path = 'D:\\上传文件\\pdf比对\\1专业数据\\数据组-研报\\2\\29'
    file_path = 'D:\\上传文件\\pdf比对\\1专业数据\\数据组-研报\\3\\31'

    # file_path = file_path.replace('\\','\\\\')
    return file_path

def judge_file():
    """ 判断PDF是研报还是公告,single"""
    file_namas = 'D:\\上传文件\\pdf比对\\复杂文件\\非研报类\\生物可降解材料的临床应用.pdf'
    path = os.path.dirname(__file__)
    url = "http://192.168.1.223:7071/convert/pdf/pdf_fragment/pdf_clf"
    files = {'file':open(file_namas,'rb')}
    try:
        res = requests.post(url=url, files=files)
        print(res.text)
    except Exception as e:
        print(e)


def sendmessage():
    """ 造点消息"""
    url = 'https://testcyprex.fir.ai/api/resource/annotation/add/'
    data = {'res_id':'n9rGO8WW6l8Na3kE','content':'@云雾 ','uids':'[32]'}
    cookie = {'fir_session_id':'oeyi6zr5u0s33vn2pp5vmogqgutgocps'}
    res = requests.post(url=url,data=data,cookies = cookie)
    print(res.text)

def get_url():
    """ 获取url"""
    # file = open("url.txt",'r+')
    # lens = len(file.readlines())
    # for i in range(40000):
    #     file.readline()
    wb = openpyxl.load_workbook('url.xlsx')
    content = wb.worksheets[0]
    count = content.max_row
    print(count)
    for i in range(2):
        url = content.cell(i+1,2).value
        print(url)
        print(os.path.split(url)[1])






def download_files(url=None):
    """ 下载文件"""
    # url = 'http://disc.static.szse.cn/download//disc/disk01/finalpage/2019-04-15/fec065d5-27f5-449d-b781-95e03d18ac46.PDF'
    # name = os.path.split(url)
    # file = open("url.txt", 'r+')
    # lens = len(file.readlines())
    # for i in range(40000):
    #     url = file.readline()

    wb = openpyxl.load_workbook('url.xlsx')
    content = wb.worksheets[0]
    count = content.max_row
    print('总共 %s 条'%count)
    j = 0
    for i in range(count):
        url = content.cell(i+1,2).value
        # name = os.path.split(url)
        name = str(os.path.split(url)[1])
        # print(name)
        try :
            res = requests.get(url)
            with open(name,"wb") as code:
                code.write(res.content)
            j+=1
            if (i + 1) % 500 == 0:
                print('进行到第 %s 条' % (i + 1))
            elif i == count - 1:
                print('执行完成，总共成功 %s 条'%j)
        except Exception as e:
            print(e)
            continue



if __name__=="__main__":
    # file_path = getUrl()
    # judge_files()
    # for i in range(35):
    #     sendmessage()
    # get_url()
    # download_files()
    pass




























