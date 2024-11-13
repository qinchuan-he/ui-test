# coding = utf-8

# 智能搜索验证,其中包含了jmeter报告的检查方法
from concurrent.futures.thread import ThreadPoolExecutor

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
from time import sleep
import time
import openpyxl
from common.comfunction import User
from common.comfunction import OpenBrowser
from common.comfunction import highlight
from common.comparameter import symbol
from common.comfunction import com_xpath
from selenium.webdriver.common.keys import Keys
from common.comfunction import send_mail
from common.private import EmailProperty,folder_path
import os
import requests
import json
from common.comparameter import symbol



class search_home(object):
    """ 智能搜索首页的相关操作，需要按照函数顺序来执行 """
    def lately_collection(self,driver,image_path=None, image_prefix=None):
        """ 最近收藏"""
        # 切换到智能搜索
        User().switch_navigation(driver, name="智能搜索")
        try:
            # 是否有最近收藏
            driver.find_element_by_xpath("//div[contains(@class,'GlobalSearchPage_searchBody')]")
            if image_path:
                driver.get_screenshot_as_file(image_path+image_prefix+"-收藏原格式"+str(time.time())+".png")

            try:
                WebDriverWait(driver,5,0.5).until(ec.element_to_be_clickable((By.XPATH,"//span[text()='纯文本']/..")))
                el1 = driver.find_element_by_xpath("//span[text()='纯文本']/..")
                highlight(driver,el1)
                el1.click()
                sleep(1)
                if image_path:
                    driver.get_screenshot_as_file(image_path+image_prefix+"-收藏原格式"+str(time.time())+".png")
            except Exception as e:
                print("未解析出纯文本")
        except Exception as e:
            print("没有收藏内容")

    def my_subscribe(self,driver,image_path=None,image_prefix=None):
        """ 我的订阅 ，前提是处于智能搜索页面"""
        # 进入我的订阅
        sleep(1)
        driver.find_element_by_xpath("//li[text()='我的订阅']").click()
        sleep(1)
        if image_path:
            driver.get_screenshot_as_file(image_path+image_prefix+"-订阅截图"+str(time.time())+".png")

    # 我的批注,前提是进入智能搜索首页
    def my_annotation(self,driver,image_path=None,image_prefix=None):
        """ 我的批注"""
        sleep(1)
        driver.find_element_by_xpath("//li[text()='我的批注']").click()
        sleep(1)
        if image_path:
            driver.get_screenshot_as_file(image_path+image_prefix+"-我的批注"+str(time.time())+".png")

class search_result(object):
    # 智能搜索
    def search(self,driver,image_path=None,image_prefix=None):
        # 切换到智能搜索
        User().switch_navigation(driver, name="智能搜索")
        el = com_xpath().smart_search(driver)
        el.send_keys("*")
        driver.switch_to.active_element.send_keys(Keys.ENTER)
        sleep(2)
        for i in symbol().english_symbol:
            el1 =  driver.find_element_by_xpath("//input[@type='text']")
            el1.send_keys(Keys.BACK_SPACE)
            el1.clear()
            el1.send_keys(i)
            driver.switch_to.active_element.send_keys(Keys.ENTER)
            sleep(3)
            if image_path:
                driver.get_screenshot_as_file(image_path+image_prefix+str(time.time())+".png")
        print("英文符号完毕")
        for i in symbol().china_symbol:
            el1 = driver.find_element_by_xpath("//input[@type='text']")
            el1.send_keys(Keys.BACK_SPACE)
            el1.clear()
            el1.send_keys(i)
            driver.switch_to.active_element.send_keys(Keys.ENTER)
            sleep(3)
            if image_path:
                driver.get_screenshot_as_file(image_path + image_prefix + str(time.time()) + ".png")
        print("中文符号完毕")

    def check_jmeter(self,driver,report_url,position,image_path=None,image_prefix=None):
        """
        检查jmeter执行情况
        :param driver:
        :param image:
        :param image_prefix:
        :return:
        """

        # 2020-06-05增加方法，创建jmeter报告截图
        jmeter_path = os.path.join(folder_path, '截图', 'jmeter报告')
        if os.path.exists(jmeter_path):
            for i in os.listdir(jmeter_path):
                os.remove(os.path.join(jmeter_path, i))
        else:
            os.mkdir(jmeter_path)

        sleep(10)
        driver.get(report_url)
        sleep(0.5)
        el = driver.find_elements_by_xpath("//div[@style='font-size:8pt;text-align:center;padding:2px;color:white;']")
        if len(el)>1:
            print("有问题")

            # 2020-06-05 增加，错误排序之后滚动
            el3s = driver.find_elements_by_xpath("//div[@class='tablesorter-header-inner']")
            screenshot = []
            screenshot_name = []
            for i in range(len(el3s)):
                if 'Error' in el3s[i].text:
                    if '%' in el3s[i].text:
                        ActionChains(driver).move_to_element(el3s[i]).perform()
                        sleep(0.5)
                        ActionChains(driver).double_click(el3s[i]).perform()
                        for loop in range(2):
                            sleep(0.5)
                            ActionChains(driver).send_keys(Keys.PAGE_DOWN).perform()
                        images_url= os.path.join(jmeter_path,str(time.time())+'.png')
                        driver.get_screenshot_as_file(images_url)
                        screenshot.append(images_url)
                        screenshot_name.append('error'+str(i)+'.png')
                        break

            if image_path:
                driver.get_screenshot_as_file(image_path+image_prefix+str(time.time())+".png")
            # 2020-06-05 增加截图
            if position==1:
                send_mail("接口检查有问题",EmailProperty().EMAIL_ATTACHMENT1,screenshot,screenshot_name)
            elif position==2:
                send_mail("接口检查有问题", EmailProperty().EMAIL_ATTACHMENT2, screenshot, screenshot_name)
        else:
           s = str(el[0].text)
           a = s.split("\n",2)[0]
           if a=='OK':
               print("全部通过")
               # send_mail("没有问题", EmailProperty().EMAIL_ATTACHMENT2, EmailProperty().EMAIL_ATTACHMENT2, "哎临时用下.html")


        # sleep(5)


# 新版本的站内搜索  2020-08-04
class site_serach_rcheck():
    cookie={'fir_session_id':'kstm0mpewpu7xux0zuzohliq95xn51cq'}

    # ,2020-08-28调整，返回文件json结果
    def get_result(self,search_keywords,search_type,clickSearch,page=None,tag_name=None,start_time=None,end_time=None,author_list=None):
        """ 获取分词"""
        if not page:
            page=1

        url = 'https://testcyprex.fir.ai/api/resource/search/'
        if start_time and end_time and author_list:
            parament = {'search_keywords': search_keywords, 'search_type': search_type, 'clickSearch': clickSearch, 'start_time': ''
                , 'end_time': '', 'page_row': '20', 'page': str(page), 'keywords_pos': '0', 'info_type': ''
                , 'ordering': 'score', 'author_list': '[]', 'is_correct': 'true'}
        else:
            parament = {'search_keywords': search_keywords, 'search_type': search_type, 'clickSearch': clickSearch,
                        'start_time': ''
                , 'end_time': '', 'page_row': '20', 'page': '1', 'keywords_pos': '0', 'info_type': ''
                , 'ordering': 'score', 'author_list': '[]', 'is_correct': 'true'}
        if tag_name:
            parament.setdefault('tag_name','')
        # print(parament)
        try:
            res = requests.get(url=url,params=parament,cookies=self.cookie)
            # print(res.text)
            # result = json.loads(res.text).get('data')
            # cut_words = result.get('cut_words')
            # similar_words = result.get('similar_words')
            # corrects = result.get('corrects')
            # # print(cut_words)
            # # print(similar_words)
            # # print(corrects)
            # return cut_words,similar_words,corrects
            return json.loads(res.text)
        except Exception as e:
            return e



# 阅读excel,上面方法改了，需要改，或者使用时候把上面注释放开
def read_excel():
    search_keywords = '中华人民共盒国'
    search_type = '002'
    clickSearch = 'false'
    tag_name = '126'
    path = r'D:\work\1测试\2用例\cypress系统\cyprex2.2.5和边写边搜插件\搜索数据.xlsx'
    wb = openpyxl.load_workbook(path)
    # for i in wb.worksheets:
    sheet = wb.worksheets[0]
    max = sheet.max_row+1
    max = len(symbol.chinese)

    for i in range(2,sheet.max_row+1):
        search_keywords = sheet.cell(i, 2).value
    # for i in range(2, max+2):
    #     search_keywords = symbol.chinese[i-2]
    #     sheet.cell(i, 2).value = str(search_keywords)
        cut_words,similar_words,corrects = site_serach_rcheck().get_result(search_keywords, search_type, clickSearch, tag_name)
        if cut_words:
            sheet.cell(i, 5).value = str(cut_words)
        if similar_words:
            sheet.cell(i, 6).value = str(similar_words)
        if corrects:
            sheet.cell(i, 3).value=corrects[0]
        print('执行中...')
    path_2 = r'D:\work\1测试\2用例\cypress系统\cyprex2.2.5和边写边搜插件\搜索数据'+time.strftime('%Y_%m_%d-%H_%M_%S',time.localtime(time.time()))+'.xlsx'
    wb.save(path_2)
    print('执行完成')

# 检查特殊符号的搜索结果,本次检查后端是否有报错
def symbolSearch():
    search_keywords = '中华人民共盒国'
    search_type = '002'
    clickSearch = 'false'
    msg_list = []
    for i in symbol.english:
        search_keywords = i
        result = site_serach_rcheck().get_result(search_keywords, search_type, clickSearch)

        try:
            # print('{}==>{}'.format(i,result.get('msg')))

            if result.get('status')==0:
                print('{}----存在异常----'.format(i))
            else:
                num=result.get('data').get('data_list')
                msg = {i: result.get('msg'),'num':len(num)}
                msg_list.append(msg)
        except Exception as e:
            print(e)
            msg = {i: e}
            msg_list.append(msg)
            continue
    print(msg_list)

def read_file_search():
    search_type = '002'
    clickSearch = 'false'
    msg_list = []
    path_2 = r'D:\work\3文档\招聘\d.txt'
    path_3 = r'D:\work\3文档\招聘\f.txt'
    with open(path_2,'r+',encoding='utf-8') as file:
        with open(path_3,'a+',encoding='utf-8') as file_3:
            while True:
                search_keywords = file.readline()
                if search_keywords:
                    try:
                        result = site_serach_rcheck().get_result(search_keywords, search_type, clickSearch)
                        if result.get('status') == 0:
                            print('{}----存在异常----'.format(search_keywords))
                            s = {search_keywords:result.get('msg')}
                            msg_list.append(s)
                            file_3.write(search_keywords)
                    except  Exception as e:
                        s = {search_keywords: '出现异常'}
                        msg_list.append(s)
                        file_3.write(search_keywords)
                        print('{}:{}'.format(e,search_keywords))

                else:
                    break
    print('over')
    print(msg_list)


# 查询返回结果是否有em标签
def search_mark():
    search_type = '002'
    clickSearch = 'false'
    search_keywords='公司，股份，进入 哇哈哈'
    try:
        result = site_serach_rcheck().get_result(search_keywords, search_type, clickSearch)
        num = result.get('data').get('page').get('total_pages')
        for i in range(1,num+1):
            print(i)
            print('---------------------------')
            result_1 = site_serach_rcheck().get_result(search_keywords, search_type, clickSearch,page=i)
            data_list=result_1.get('data').get('data_list')
            for j in data_list:
                s=j.get('norm_content')
                if '<' in s or '>' in s:
                    print('{}:{}:{}'.format(j.get('oid'),j.get('name'),s))
    except Exception as e:
        print(e)


# 检查公告实体，目前实体对应公司，通过公司股票代码，简称，全程去对应
def check_entity(key):
    cookie = {'fir_session_id':'czcoaqfce9mgwsipavzr2j5d1k2cfqvl'}
    if type(key)==str:
        key='国泰君安'
        url = 'https://testcyprex.fir.ai/api/resource/publicSearch/?url=%2Fresource%2FpublicSearch%2F&' \
              'search_keywords=' + key + '&table_code=004&info_type=02&ordering=score' \
                                         '&search_level=1&start_time=&end_time=&page_row=20&page=2&author_list=%5B%5D&is_correct=true'
        res = requests.get(url=url, cookies=cookie)
        # print(res.text)
        expand_list = json.loads(res.text).get('data').get('results')[0].get('expand')
        print(expand_list)
    else:
        for i in key:
            url = 'https://testcyprex.fir.ai/api/resource/publicSearch/?url=%2Fresource%2FpublicSearch%2F&' \
                  'search_keywords='+i+'&table_code=004&info_type=02&ordering=score' \
                  '&search_level=1&start_time=&end_time=&page_row=20&page=2&author_list=%5B%5D&is_correct=true'
            res = requests.get(url=url,cookies=cookie)
            # print(res.text)
            expand_list=json.loads(res.text).get('data').get('results')[0].get('expand')
            # print(expand_list)
            if i not in expand_list:
                print('{}:命中实体有问题'.format(i))
def aac():
    # 读取Excel文件
    path = os.path.join(r'D:\work\1测试\16测试数据\entity.xlsx')
    cookie = {'fir_session_id': 'czcoaqfce9mgwsipavzr2j5d1k2cfqvl'}
    wk = openpyxl.load_workbook(path)
    key_list = []
    for i in wk.worksheets:
        rows = i.max_row
        for j in range(rows):
            cell = i.cell(j+1,1).value
            key = cell.split(',')[0]
            # print(type(cell))
            # print(cell.split(',')[0])
            #
            # url = 'https://testcyprex.fir.ai/api/resource/publicSearch/?url=%2Fresource%2FpublicSearch%2F&' \
            #       'search_keywords='+key+'&table_code=004&info_type=02&ordering=score' \
            #       '&search_level=1&start_time=&end_time=&page_row=20&page=2&author_list=%5B%5D&is_correct=true'
            # res = requests.get(url=url, cookies=cookie)
            # # print(res.text)
            # expand_list = json.loads(res.text).get('data').get('results')[0].get('expand')
            # if key not in expand_list:
            #     print('{}拓展有问题'.format(key))

            # print(expand_list)
            key_list.append(key)
    return key_list

def threads_entity():
    # 实体开启多线程
    files = aac()
    threads = ThreadPoolExecutor(max_workers=6)
    file_s = []
    num = int(len(files) / 6)
    files_1 = files[0:num]
    files_2 = files[num:2 * num]
    files_3 = files[2 * num:3 * num]
    files_4 = files[3 * num:4 * num]
    files_5 = files[4 * num:5 * num]
    files_6 = files[5 * num:]
    file_s.append(files_1)
    file_s.append(files_2)
    file_s.append(files_3)
    file_s.append(files_4)
    file_s.append(files_5)
    file_s.append(files_6)
    # 线程池启动
    for i in range(6):
        future=threads.submit(check_entity,file_s[i])
        print('启动线程{}'.format(i))
    threads.shutdown(wait=True)











if __name__=='__main__':
    # search_keywords='国泰军安'
    search_keywords = '中华人民共盒国'
    search_type='002'
    clickSearch='false'
    tag_name='126'
    path = r'D:\work\1测试\2用例\cypress系统\cyprex2.2.5\搜索数据.xlsx'
    # site_serach_rcheck().get_result(search_keywords,search_type,clickSearch)
    # read_excel()
    # symbolSearch()
    # read_file_search()
    # search_mark()
    # check_entity()
    # aac()
    threads_entity()














