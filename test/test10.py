# coding=utf-8
import requests
import os
import json
import time
import openpyxl
import datetime
import random

def json_check():
    s1='{"asctime":"﻿2021-03-12 14:49:44.124","thread":140081061783296,"request_id":"a725569658a648059c630ace0520aaa6","lineno":303,"module":"enter_log","level":"INFO","msg":"{"remote_ip":"﻿101.81.120.173","type":"request","method":"GET","path":"/account/user/info/","user":17702155559, "userId":﻿1444, "get_params":{"access_token": ["DPAO7NZT3CQ1BYSW0VFR865GHK4LJE2I"], "user_open_id": ["1443"]},"post_params":{}}"} '
    s1='{"asctime":"2021-03-12 14:42:08.223","thread":140081061783296,"request_id":"a725569658a648059c630ace0520aaa6","lineno":303,"module":"enter_log","level":"INFO","msg":"{"remote_ip":"172.17.0.19999","type":"request","method":"GET","path":"/account/user/info/","user":10012320171, "userId":1443, "get_params":{"access_token": ["DPAO7NZT3CQ1BYSW0VFR865GHK4LJE2I"], "user_open_id": ["1443"]},"post_params":{}}"} '
    t=json.loads(s1)
    print('---8')
    print(t)


def check_log():
    path = r'D:\work\1测试\16测试数据\loginlog.xlsx'
    wb = openpyxl.load_workbook(path)
    sheet = wb.worksheets
    max_row = sheet[0].max_row
    user_list = []
    print(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time())))
    print(time.time())
    num=0
    for i in range(1+max_row):
        u_id = sheet[0].cell(i+1,1).value
        time_s=sheet[0].cell(i+1,2).value
        time_lst=str(time_s).split(',')

        if len(time_lst)<2:
            # print(time_lst)
            continue
        for i in time_s.split(',')[1:]:
            # print(i)
            num+=1
            t = time.strptime(i.split(' ',2)[0],'%Y-%m-%d')
            if t>=time.strptime('2021-03-15','%Y-%m-%d') and t<time.strptime('2021-03-18','%Y-%m-%d'):
                if u_id not in user_list:
                    user_list.append(u_id)
                break

    print(user_list)
    print(len(user_list))
    print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))
    print(time.time())
    print(num)

def cc():
    order_id = "{}{:0>10}{:0>2}".format(datetime.datetime.now().strftime('%Y%m%d%H%M%S'),
                                        1213132,
                                        random.randint(0, 99))
    order_id2 = "{}{}{}".format(datetime.datetime.now().strftime('%Y%m%d%H%M%S'),
                                        1213132,
                                        random.randint(0, 99))
    order_id3 = "{:0>6}".format(2012555)
    print(order_id)
    print(order_id2)
    print(order_id3)




if __name__=="__main__":
    # json_check()
    cc()






















