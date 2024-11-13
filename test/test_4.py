# coding = utf-8

import os
# import time
# from time import sleep
# from common.comfunction import OpenBrowser,User,com_path
# from test.smart_search import search_home,search_result
# from buttonFunction.function_checkcontract import check_proofreading
# from common.comfunction import Image_Processing

# from PIL import Image
# from PIL import ImageChops
# import xlrd
# import xlwt
import shutil
# import pandas
import openpyxl

# import openpyxl.cell.cell
import sys
# from robotide import main
# from robotide import run
# print(sys.path)

from common.comfunction import *
from selenium.webdriver.common.keys import Keys
from buttonFunction.function_contractrelated import contratc_split
from buttonFunction.function_contractrelated import contract_combine
import json
from common.private import UserProperty


import requests

# r = requests.post('https://testcyprexsvc.fir.ai/account/user/judge/register/',data={'key':'mobile','value':'13245698565'})
# print(r.text)

# 传入域名、接口地址、方法类型、参数、是否需要携带cookie
def request_meth(system =None, url=None,method=None,params=None):
    if system == None:
        system=UserProperty().SYSTEM
    if url :
        url = system+url
        resp = ''
        if method=='POST':
            if params:
                resp = requests.post(url,data=eval(params))
            else:
                resp = requests.post(url)
        elif method == 'GET':
            if params:
                resp = requests.get(url,data=eval(params))
            else:
                resp = requests.get(url)
        else:
            print("异常")
        return resp


fp = "D:\\122.xlsx"
shutil.copyfile(fp,'D:\\1222.xlsx')
wb = openpyxl.load_workbook("D:\\1222.xlsx")
# tab = wb['我重命名的']
# print(tab.max_row)

for i in wb.sheetnames:
    for j in range(2,wb[i].max_row+1):
        lt = []
        for k in range(1,wb[i].max_column):
            # print(wb[i].cell(j,k).value)
            lt.append(wb[i].cell(j,k).value)
        # print(lt)
        # print(lt[3:6])
        res = request_meth(lt[3],lt[4],lt[5],lt[6])
        # print(res.status_code)
        print(res.text)
        print(type(res.text))
        # print(json.loads(res.text))
        # print(eval(res.text))
        # print(type(eval(res.json)))
        res_json = json.loads(res.text)
        if res_json.get('status')==1:
            wb[i].cell(j, 9).value = res.status_code
            wb[i].cell(j,10).value = 1
            wb[i].cell(j, 11).value = res_json.get('msg')
            wb[i].cell(j, 12).value = '通过'
            wb.save('D:\\1222.xlsx')
        else:
            wb[i].cell(j, 11).value = '失败'
            wb.save('D:\\1222.xlsx')
        print('执行完成')

# print(lt)
# print(wb.sheetnames)























a1 = (100+1200)*(1.1+1.06)*(50+1)
a2 = (100+1200)*(1.1+1.06)*(50+10)
a3 = (100+1200)*(1.1+1.06)*(50+50)
a4 = (100+1200)*(1.1+1.06)*(50+200)
a5 = (100+1200)*(1.1+1.06)*(50+600)
a6 = (100+1200)*(1.1+1.06)*(50+1500)
a7 = (100+1200)*(1.1+1.06)*(50+3000)
a_6 = (100+2000+2434)*(1.1+1.07+1.5)*(50+1500)/10000
a_7 = (100+2000+2434)*(1.1+1.07+1.5)*(50+3000)/10000

a_x = (100+4500+15624)*(1.1+3.4+2.9)*(50+56000)/10000/10000
a_x1 = (100+4500+10021)*(1.1+3.4+2.6)*(50+72000)/10000/10000
a_y = (100+4500+15624)*(1.1+3.4+2.9)*(50+72000)/10000/10000
a_y1 = (100+4500+15624)*(1.1+3.4+2.9)*(50+88000)/10000/10000
# print(a_y-a_x)
# print(a_y1-a_y)
#
# print(int(a_6))
#
# print(int(a_7))


a_list1 = [a1,a2,a3,a4,a5,a6,a7]
a_list = []
for i in range(7):
    a_list.append('%.2f'%(a_list1[i]/10000))
# for i in a_list1:
    # print("%.2f" %(i/10000))

for i in range(len(a_list)):
    if i<len(a_list)-1:
        s1 = float(a_list[i])
        s2 = float(a_list[i+1])
        # print('第%d级别和第%d级别倍数差距：%.2f'%(i+2,i+1,s2/s1))
print('-----------------------')
for i in range(len(a_list)):
    s1 = float(a_list[i])
    s2 = float(a_list[-1])
    # print('第%d级别和第%d级别倍数差距：%.2f' % (7, i + 1, s2 / s1))



# fp="D:\\121.xlsx"
#
# shutil.copyfile(fp,"D:\\1212.xlsx")
#
# # f = pandas.read_excel(fp,sheet_name=0)
# # print(f.head())
#
# ws = openpyxl.load_workbook("D:\\1212.xlsx")
#
# ws2 = ws.active
# # ws2 = ws['Sheet1']
# s = ws2['A1':'M5'] #type:tuple
#
# print(s[0])
# for i in s:
#     for j in i:
#         print(j.value)
# print(ws.active)
# print(ws2['D1'].value)
# ws.active.title = '我重命名的'
# print(ws.active)
# print(ws.active['D3'].value)
# print(ws.sheetnames)
# sheet = ws['Sheet1']
# num = 3
# posi = 'D'
# position = posi+str(num)
# s = sheet[position] #type:openpyxl.cell.cell.Cell
# p = s.value
# print(p)
# sheet['D3'] = '哇哈哈'
# print(sheet['D3'].value)
# ws.save("D:\\1212.xlsx")





# wk = xlwt.Workbook()
# ws = wk.add_sheet('test') #type:xlwt.Worksheet
# ws.write(0,0,'你好哇啊哈哈哈哈哈')
# wk.save('D:\\2.xls')
# print(type(ws))
# a = ['1','2','3']
# print(type(a))






# fp = xlrd.open_workbook('D:\\121.xlsx')
# # fp2 = xlrd.count_records('D:\\121.xlsx')
# # fp = open("D:\\1.txt")
#
# print(fp.sheet_by_index(0).row_values(1,1,5))
# print("---")
# # print(fp.read())















