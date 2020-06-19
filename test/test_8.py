#coding=utf-8

from common.comfunction import send_mail
import os
from common.comfunction import OpenBrowser
from time import sleep
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from common.private import EmailProperty,folder_path,DB
import shutil
import pymysql
import openpyxl
from typing import List
import requests

# sendemail

def sendemail1():
    """ send email test"""
    file_url = r'D:\work\1测试\6接口\3服务器检查脚本\report\index.html'
    image_path = r'D:\上传文件\pdf比对\1专业数据\新建文件夹'
    image_path = os.path.join(image_path)
    email_subject = 'test email'
    url = 'http://192.168.1.223:8077/jmeter/report2/index.html'
    url = 'D:/work/1测试/6接口/3服务器检查脚本/report2/index.html'
    # send_mail()



    # 启动一个浏览器
    driver = OpenBrowser(mode=2)
    driver.get(url)
    sleep(0.5)
    el = driver.find_elements_by_xpath("//div[@style='font-size:8pt;text-align:center;padding:2px;color:white;']")

    # print('准备截图')
    # els = driver.find_elements_by_xpath("//div[@class='panel-body']")
    # for i in range(len(els)):
    #     ActionChains(driver).move_to_element(els[i]).perform()
    #     image = driver.get_screenshot_as_file(image_path + '\\' + str(time.time()) + '.png')
    # print('截图完成')

    el3s = driver.find_elements_by_xpath("//div[@class='tablesorter-header-inner']")
    for i in range(len(el3s)):
        if 'Error' in el3s[i].text:
            if '%' in el3s[i].text:
                ActionChains(driver).move_to_element(el3s[i]).perform()
                ActionChains(driver).double_click(el3s[i]).perform()
                ActionChains(driver).send_keys(Keys.PAGE_DOWN).perform()




    print(len(el3s))
    el2s = driver.find_elements_by_xpath("//tr[@role='row']/td[3]")
    # for i in range(len(el2s)):
    #     print(el2s[i].text)
    sleep(1)
    driver.quit()

def yousee():
    a = [1,8,6,2,5,4,8,3,7]
    area_max = 0
    i = 0
    j = len(a)-1

    area_max=0
    while(j!=i):
        if area_max<(j-i)*min(a[i],a[j]):
            area_max=(j-i)*min(a[i],a[j])
        if a[i]>a[j]:
            j-=1
        else:
            i+=1
    print(area_max)

def insert_content(parameter:List):
    """ 往数据库插入东西"""
    url = r'D:\5python'
    host1 = DB.host
    port1 = DB.port
    user1 = DB.user
    password1 = DB.password
    db_name = DB.db_name
    connect = pymysql.connect(host =host1,port = int(port1),user = user1,password=password1,database='test',charset='utf8')
    cur = connect.cursor()
    # s = cur.execute('select * from consumer')
    #     # print(cur.fetchall())
    try:
        sql = "insert into ebook_name(path,`name`) values(%s,%s);"
        print(sql)
        # ls = [('4','3','4'),('3','3','4')]
        # cur.execute("insert into ebook_name(path,`name`,`desc`) values('2','3','4');") # 单条方法
        cur.executemany(sql,parameter) # 批量方法
        connect.commit()
    except Exception as e:
        connect.rollback()
        print('出现异常，终止')
        print(e)
    print('执行完毕')

def read_excel(position=None):
    """ 读取Excel,position有值就全部循环"""
    url = r'D:\1212.xlsx'
    url = r'D:\5python\800G小说目录.xlsx'
    url = r'D:\work\1测试\1需求\cyprex2.1.9'
    wk = openpyxl.open(url)
    all = []
    for i in wk.worksheets:
        count = i.max_row
        for j in range(count):
            path = i.cell(j+1,1).value
            name = i.cell(j+1,2).value
            ll = (path,name)
            all.append(ll)
            print('-')
            # if j == 20:
            #     break

        # if not position:
        #     break
    return all

import hashlib
import json
import base64
def generate_jwt(**kwargs):
    """调试jwt的测试接口"""
    user_open_id = kwargs.get("user_open_id", "1")
    username = kwargs.get("username", "滴滴")

    p = hashlib.sha256('MzY1Mjk0OTM5NTgz'.encode("utf-8"))
    header = {
        "alg": "HS256", # 加密方式为HMAC SHA256
        "typ": "JWT"
    }
    header = base64.urlsafe_b64encode(json.dumps(header).encode('utf-8')).decode('utf-8')
    payload = {
        "iat": time.time(),  # 加密方式为HMAC SHA256
        "exp": time.time() + 60 * 60,
        "user_open_id": user_open_id,
        "username": username,
    }
    payload = base64.urlsafe_b64encode(json.dumps(payload).encode('utf-8')).decode('utf-8')
    p.update(r"{}\.{}".format(header, payload).encode("utf-8"))
    signature = p.hexdigest()

    return "Bearer {}.{}.{}".format(header, payload, signature)

def interface(head_key,resource_id=None):
    """  紫光云接口"""
    header = {"Authorization":head_key}
    print(header)
    cookie = {'fir_session_id':'vkbbijcvul6y8hw7kkzvoqp2jvbw9bv2'}



    # 创建会议接口
    # url = 'https://testcloud.fir.ai/api/group/meeting/create/'
    # url = 'https://devcloud.fir.ai/api/group/meeting/create/'
    # url = 'http://117.9.200.62:20080/api/group/meeting/create/'
    # parameters = {'name':'第一个IM','team_type':'3','team_open_id':'1','extra_info':'来自接口'}
    # res = requests.post(url=url,data=parameters,headers=header)
    # print(res.text)

    # 添加会议成员
    # url='https://testcloud.fir.ai/api/group/meeting/member/create/'
    # pp = [{'user_open_id':'1','user_name':'气喘','name':'hahah'},{'user_open_id':'33','user_name':'气喘','name':'hahah'}]
    # parameter={'team_type':'3','team_open_id':'222224','user_open_id_list':json.dumps(pp)}
    # res = requests.post(url=url,data=parameter,headers = header)
    # print(res.text)

    # s上传会议文件
    # url = 'https://testcloud.fir.ai/api/group/meeting/resource/upload/'
    # url = 'http://117.9.200.62:20080/api/group/meeting/resource/upload/'
    # file_path =os.path.join(r'D:\上传文件\自动化验证文档\19种格式\office')
    # # file_path = os.path.join(r'D:\上传文件\自动化验证文档\19种格式\图例提取文件')
    # # file_path = os.path.join(r'D:\上传文件\自动化验证文档\19种格式\合并')
    # # file_path = os.path.join(r'D:\上传文件\自动化验证文档\19种格式\比对文件')
    # count = 1
    # for i in os.listdir(file_path):
    #     print(i)
    #     file_info = {'name':i,'author':'qinchuan'}
    #     file={'file':open(os.path.join(file_path,i),'rb')}
    #     date = {'team_type':'2','team_open_id':'1','resource_id':count,'file_info':json.dumps(file_info)}
    #     res = requests.post(url=url,data=date,headers = header,files=file)
    #     print(res.text)
    #     print(count)
    #     count+=1

    # 文件重命名
    url = 'https://testcloud.fir.ai/api/group/meeting/resource/update/'
    date = {'resource_id':'50','name':'单一名字.pdf'}
    res = requests.post(url=url,data=date,headers=header)
    print(res.text)

    # 文件删除
    # url='https://testcloud.fir.ai/api/group/meeting/resource/delete/'
    # data = {'resource_id':'50'}
    # res = requests.post(url=url,data=data,headers=header)
    # print(res.text)




    # 接口
    # url = 'https://testcloud.fir.ai/api/resource/data/fetch/'
    # resource_id = 'AxrEkReGG4rP52BY'
    # request_url=url+'?id='+resource_id+'&view_type=preview'
    # print(request_url)
    # url_s='https://testcloud.fir.ai/api/resource/data/fetch/?id=AxrEkReGG4rP52BY&view_type=preview'
    # res = requests.get(url=request_url,headers = header)
    # print(res.text)



if __name__=='__main__':
    # sendemail1()
    # yousee()
    # insert_content()
    # result = read_excel()
    # insert_content(result)
    key = generate_jwt()
    # print(key)
    interface(key)









































