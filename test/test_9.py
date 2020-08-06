# -*- coding: utf-8 -*-
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import os
import requests
from common.lenovo import gen
import openpyxl
import json

# 批量任务线程池版
def batch_processing(task_func, args_list): # 参数: 函数名字 函数所需参数
    if len(args_list) == 1:
        return [task_func(*args_list[0])]
    else:
        results = []
        with ThreadPoolExecutor(max_workers=5) as executor:  # 线程池
            future_set = set()
            for args in args_list:
                future = executor.submit(task_func, *args)
                future_set.add(future)

            for future in as_completed(future_set):  # 阻塞等待所有任务完成
                try:
                    result = future.result()
                except Exception as e:
                    print(e)
                else:
                    results.append(result)
            executor.shutdown()
        print(results)
        return results



def one():
    cookie = {'X-LENOVO-SESS-ID': 'a0t18ie625pj4jptsemigobno1'}
    url = "http://192.168.1.224:8061/api/file/parse?"
    file_id = '90692'
    rev='dc48275c23634a78b8dd82d161de9698'
    params = {"file_id": file_id, "rev": rev, "opt_type": "01"}
    url_3 = gen(url, params)
    res = requests.get(url_3)
    print(res.text,res.status_code)

def read_excel():
    rest = []
    wk = openpyxl.load_workbook('lenovo.xlsx')
    wb = wk.worksheets[0]
    print(wb.cell(2,2).value)
    row = wb.max_row
    for i in range(2,row+1):
        a = {'rev':wb.cell(i,2).value,'neid':wb.cell(i,3).value}
        rest.append(a)
        # print(a)
    return  rest

def create_url(lst):
    url = 'http://192.168.1.224:8061/api/file/parse?'
    url = 'api/file/parse?'
    file = open(r'D:\work\1测试\16测试数据\1.txt', 'w+')
    for i in lst:
        params = {"file_id": i.get('neid'), "rev": i.get('rev'), "opt_type": "01"}
        url_2 = gen(url, params)
        # print(url_2)
        file.write(url_2+str('\n'))


    file.close()

def delete_folder():
    url = 'http://192.168.1.224:8061/api/file/parse?'
    file_id = '/hahhah/'
    rev = ''
    params = {"file_id": file_id, "rev": rev, "opt_type": "05"}
    url_2 = gen(url, params)
    res = requests.get(url_2)
    print(res.text)

#单个数据造索引
def single_indexes():
    url = 'http://192.168.1.224:8061/api/file/parse?'
    neid = '94247'
    rev = '1c326222eea644819142cd1245431138'
    params = {"file_id": neid, "rev": rev, "opt_type": "01"}
    url_2 = gen(url, params)
    res = requests.get(url_2)
    print(res.text)

# 造预览调用接口url
def create_url2(lst):
    url = 'http://192.168.1.224:8060/data/fetch?'
    url =  '/api/data/fetch?'
    file = open(r'D:\work\1测试\16测试数据\preview.txt','w+')
    for i in lst:
        params = {"view_type":"preview","file_id": i.get('neid'), "file_name":"cesces.pdf","rev": i.get('rev'),"user_id":27}
        url_2 = gen(url,params)
        file.write(url_2+str('\n'))
    file.close()


def insert_doc(main_file,meterial_file,result_file):
    url = 'http://192.168.1.223:8001/fp/extract/word/sdtInsert'
    head = {'Content-Type':'multipart/form-data'}
    file = {'file':open(main_file,'rb')}
    with open(meterial_file, 'r', encoding='utf8') as f:
        data_s = f.readline()
    print(data_s)
    data1 = {'source_details':data_s}
    # data1 = {'source_details':'{"纯文本控件-1": {"content": "xxxxx公司"},"纯文本控件-1": {"content": "肯打鸡集团"}}'}
    start = time.time()
    res = requests.post(url=url,data=data1, files=file,)
    if res.status_code==200:
        with open(result_file, "wb") as f:
            f.write(res.content)
        print('控件插入完成')
    else:
        print(res.text)
        print('error')
    print(time.time() - start)

def create_parameter(num,file_name):
    p = {"纯文本控件-0": {"content": "xxxxx公司"}}
    with open(r'D:\work\1测试\9性能+安全性\数据联动\2.txt','r',encoding='utf8') as f:
        for i in range(1,num):
            p.setdefault("纯文本控件-"+str(i),{"content": f.readline()+f.readline()})
            # print(p)
    s = str(p).replace('\'','"')
    with open(r'D:\work\1测试\9性能+安全性\数据联动\\'+file_name,'w',encoding='utf8') as f:
        f.write(s)
    print('创建完毕')

#登录
def login(mobile = None):
    url = 'https://testcyprex.fir.ai/api/account/user/signin/'
    if mobile is None:
        mobile = '10023233232'
    print(mobile)
    data = {'type': 'account','username_no': mobile,'passwd': 'Test123456','validCode': '','inviteCode': '','userId': '','teamId': ''}
    res = requests.post(url=url,data=data)
    print(str(res.headers.get('Set-Cookie')).split(';')[0].split('=')[1])
    return str(res.headers.get('Set-Cookie')).split(';')[0].split('=')[1]

def invite(session=None):
    id = 1
    url = 'https://testcyprex.fir.ai/api/group/team/invite/validation/?inviteCode=22572b92-35bf-4908-838f-01d725680139&userId='+str(id)+'&teamId=AkX3KGRaedRwpYZj'
    head = {'fir_session_id':session}
    parameter = {}
    res = requests.get(url,cookies = head)
    print(res.text)

def cs():
    url = 'https://192.168.1.224:8074/api/account/user/judge/register/'
    url = 'http://cyprexplugsvc.fir.ai/account/user/info/'
    data = {'key':'mobile','value':'10025253653'}
    res = requests.post(url=url,data=data)
    res = requests.get(url=url)
    print(res.text)




if __name__=='__main__':
    # one()
    # res = read_excel()
    # a = [{'rev':'dc48275c23634a78b8dd82d161de9698','neid':'90692'},{'rev':'711d019455dd40359ab6eeff062df2fb','neid':'90687'}]
    # create_url(res)
    # create_url2(res)
    # single_indexes()
    # delete_folder()
    # main_file = r'D:\上传文件\office文件\300控件文档\4个控件\数据联动样本.docx'
    #
    # result_file = r'D:\上传文件\office文件\300控件文档\800个控件\800_result1.docx'
    # num = 4
    # file_name = '800.txt'
    # meterial_file = os.path.join(r'D:\work\1测试\9性能+安全性\数据联动',file_name)
    # insert_doc(main_file,meterial_file,result_file)
    # # create_parameter(num,file_name)
    # # write_txt()
    # read_txt()
    # mobs = ['10101010101','12121212121','10052365214','10025637185','10052365236','10022145489','10022145488','10022145487','10025696355','10067845911','10058652325','10011111111','13818652232','10052632542','10052638541','10025253635','10056966528','17091921573','10056966525','10025252365','10058695632','10024934501','15221728696','10025253638','10025253639','17621371409','18521703809','13248131618','10058585555','10012345677','10011111113','10011111114','13131313131','10034345658','10034345659','15026835870','10023652214','15586371464','10025369652','10025369653','10025369654','10012365963','10012121345','10023233232','10015161718','15800728714','17688810823','10025699635','10058585558','10025699852','10012344444','10021248089','17778053810','30019855867','10025369657','18616060910','10022521222','19821230240','17702155559','10058696526']
    # for mob in mobs:
    #     session = login(mob)
    #     invite(session)
    cs()




